from XPdf import XPdf
from datetime import datetime
from openpyxl import load_workbook

def create_sheets(workbook, my_pdf):
    for key, value in my_pdf.files_per_month_per_year.items():
        sheet_title = "scrap_20"+str(key)
        if sheet_title not in workbook.sheetnames :
            ws = workbook.create_sheet(sheet_title)

def put_date(ws, value, rowx, columny) :      
	if (ws.cell(row = rowx, column=columny).value) == None:
		ws.cell(row = rowx, column=columny).value = value
	else:
		ws.cell(row = rowx, column=columny).value += value

def is_date_superior(key_date, table_date) :
    if isinstance(key_date, str):
         key_date = datetime.strptime(key_date, "%d/%m/%y")
    if isinstance(table_date, str):
         table_date = datetime.strptime(table_date, "%d/%m/%y")
    if key_date > table_date:
         return True
    else:
         return False
     
def get_row(ws, key):
    rowx=0
    for cell in ws['A']:
        if cell.value == key:
            rowx = cell.row
            return (rowx)
    for cell in ws['A']:
         if cell.row > 1 and cell.value != None and not is_date_superior(key, cell.value):
            rowx = cell.row
            ws.insert_rows(rowx)
            ws.cell(row=rowx, column=1).value = key
            return rowx
    for col in ws.iter_cols(min_col=1, max_col=1):
        last_cell = None
        for cell in col:
            if cell.value is not None:
                last_cell = cell
    if last_cell == None:
        rowx = 2
    else  :
        rowx = last_cell.row + 1
    ws.insert_rows(rowx)
    ws.cell(row=rowx, column=1).value = key
    return rowx
        
def get_column(ws, bill_date):
    for cell in ws[1] :
        if(cell.value == bill_date) :
            column = cell.column
            return (column)
    ws.insert_cols(2)
    ws.cell(row=1, column=2).value = bill_date
    return 2

def put_dates_per_year_sheet(workbook, my_pdf):
    create_sheets(workbook, my_pdf)
    for key, value in my_pdf.files_per_date.items() :
        year = (key[6:8])
        month = int(key[3:5]) - 1
        sheet_title = "scrap_20"+year
        ws = workbook[sheet_title]
        column = get_column(ws, my_pdf.bill_date)
        row = get_row(ws, key)
        put_date(ws, value, row, column)

def prepare_resume_sheet(ws):
    ws['B1'] = "total_files"
    ws['C1'] = "net_revenue"

def is_bill_already_scraped(workbook, date):
    try :
        ws = workbook['pdf_scraped']
    except KeyError :
        ws = workbook.create_sheet('pdf_scraped')
        prepare_resume_sheet(ws)
    for cell in ws['A']:
        if cell.value == date:
            return True
    ws.insert_rows(2)
    ws['A2'] = date
    return False

def put_rest_of_data(workbook,my_pdf):
    rowx = 0
    ws = workbook['pdf_scraped']
    for cell in ws['A']:
        if cell.value == my_pdf.bill_date:  
            rowx= cell.row
    ws.cell(row=rowx, column = 2).value = my_pdf.total_files
    ws.cell(row=rowx, column = 3).value = my_pdf.net_revenue
    
def parsing_to_excel(inputxlsx, my_pdf):
    workbook = load_workbook(inputxlsx)
    if is_bill_already_scraped(workbook, my_pdf.bill_date) == True :
        return("Error : pdf already scraped")
    put_dates_per_year_sheet(workbook, my_pdf)
    put_rest_of_data(workbook, my_pdf)
    workbook.save(inputxlsx)
    return ("Data extraction and saving successful")



